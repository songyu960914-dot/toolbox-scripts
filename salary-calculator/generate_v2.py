# -*- coding: utf-8 -*-
"""生成 run_salary_v2.py"""
import os

base_dir = r'C:\Users\yuson\.openclaw\workspace\salary-calculator'
own_b64 = open(os.path.join(base_dir, 'own_template_b64.txt')).read().strip()  # 需要本地生成
out_b64 = open(os.path.join(base_dir, 'outsource_template_b64.txt')).read().strip()  # 需要本地生成

script = r'''# -*- coding: utf-8 -*-
"""
工资计算工作流 v2 - 内嵌模板版本

用法：
  python run_salary_v2.py <总文件路径>                    # 模式1：使用内嵌模板
  python run_salary_v2.py <总文件> <自有表> <外包表>       # 模式2：使用自定义表

交互功能：运行后会询问是否需要配置负责人-下属关系进行数据汇总。
支持两种方式：1. 交互式手动输入  2. 从Excel文件读取
"""

import os
import sys
import base64
import tempfile
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# ============================================================
# 内嵌模板（base64编码）
# ============================================================
OWN_TEMPLATE_B64 = """
''' + own_b64 + r'''
"""

OUTSOURCE_TEMPLATE_B64 = """
''' + out_b64 + r'''
"""


# ============================================================
# 配置区
# ============================================================
CONFIG = {
    # 输出目录（桌面）
    "output_dir": os.path.join(os.path.expanduser("~"), "Desktop", "工资计算结果"),
    
    # 总文件结构参数
    "total": {
        "unit_col": 1,
        "name_col": 4,
        "data_start_col": 7,
        "data_start_row": 5,
        "filter_unit": "平舆",
    },
    
    # 自有表结构参数（模板）
    "own": {
        "sheets": {
            "县城自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
            "乡镇自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
        }
    },
    
    # 外包表结构参数（模板）
    "outsource": {
        "sheets": {
            "县城外包": {
                "name_col": 3,
                "data_start_row": 6,
                "data_start_col": 6,
            },
            "乡邮外包": {
                "name_col": 4,
                "data_start_row": 6,
                "data_start_col": 7,
            },
        }
    },
    
    # 使用自定义文件时的结构参数
    "own_custom": {
        "sheets": {
            "县城自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
            "乡镇自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
        }
    },
    "outsource_custom": {
        "sheets": {
            "县城外包": {
                "name_col": 3,
                "data_start_row": 6,
                "data_start_col": 6,
            },
            "乡邮外包 ": {
                "name_col": 4,
                "data_start_row": 6,
                "data_start_col": 7,
            },
        }
    },
}

# 硬编码列映射规则
COLUMN_MAPPING_RULES = {
    "default": {
        **{i: i - 7 for i in range(7, 32)},
        32: 25,
        33: 26,
        34: 27,
        # col 35: 跳过
        36: 28,
        37: 29,
        38: 30,
        39: 31,
        40: 32,
        41: 33,
    },
}


def decode_template(b64_str):
    """解码 base64 模板为临时文件路径"""
    data = base64.b64decode(b64_str.strip())
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.write(data)
    tmp.close()
    return tmp.name


def load_total_file(total_file, config):
    """读取总文件，筛选平舆人员"""
    total_cfg = config["total"]
    df = pd.read_excel(total_file, header=None)
    
    mask = df.iloc[:, total_cfg["unit_col"]] == total_cfg["filter_unit"]
    filtered = df[mask]
    
    name_to_row = {}
    for idx, row in filtered.iterrows():
        name = row.iloc[total_cfg["name_col"]]
        if pd.notna(name):
            name_to_row[str(name).strip()] = row.tolist()
    
    print(f"总文件筛选出 {total_cfg['filter_unit']} 人员: {len(name_to_row)} 人")
    return name_to_row


def paste_data(wb, sheet_name, sheet_cfg, name_to_row, mapping_rule):
    """粘贴数据"""
    ws = wb[sheet_name]
    name_col = sheet_cfg["name_col"]
    data_start_col = sheet_cfg["data_start_col"]
    data_start_row = sheet_cfg["data_start_row"]
    
    matched = 0
    not_found = []
    
    for row_idx in range(data_start_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=name_col + 1).value
        if cell_value is None:
            continue
        
        name = str(cell_value).strip()
        if name in name_to_row:
            row_data = name_to_row[name]
            for total_col_idx, target_offset in mapping_rule.items():
                if total_col_idx < len(row_data):
                    val = row_data[total_col_idx]
                    if pd.notna(val):
                        target_col = data_start_col + 1 + target_offset
                        ws.cell(row=row_idx, column=target_col, value=val)
            matched += 1
        else:
            not_found.append(name)
    
    print(f"  [{sheet_name}] 匹配成功: {matched} 人", end="")
    if not_found:
        print(f", 未在总文件中找到: {not_found}")
    else:
        print()
    
    return matched


def input_manager_relations_interactive():
    """交互式输入负责人-下属关系"""
    print()
    print("-" * 40)
    print("手动输入负责人-下属关系")
    print("输入格式: 先输入负责人姓名，再输入下属姓名（逗号分隔）")
    print("输入空行结束")
    print("-" * 40)
    print()
    
    relations = {}
    
    while True:
        manager = input("负责人姓名（空行结束）: ").strip()
        if not manager:
            break
        
        subordinates_str = input(f"  {manager} 的下属（逗号分隔）: ").strip()
        if not subordinates_str:
            print("  跳过（无下属）")
            continue
        
        subordinates = [s.strip() for s in subordinates_str.split(",") if s.strip()]
        if subordinates:
            relations[manager] = subordinates
            print(f"  已记录: {manager} <- {subordinates}")
        print()
    
    return relations


def input_manager_relations_excel():
    """从Excel文件读取负责人-下属关系
    
    Excel格式要求：
    - 第1列：负责人姓名
    - 第2列：下属姓名（同一负责人多个下属占多行，或逗号分隔）
    """
    print()
    print("-" * 40)
    print("从Excel文件读取负责人-下属关系")
    print("文件格式: 第1列=负责人, 第2列=下属（逗号分隔或多行）")
    print("-" * 40)
    print()
    
    file_path = input("请输入Excel文件路径: ").strip().strip('"')
    
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在 - {file_path}")
        return {}
    
    df = pd.read_excel(file_path, header=None)
    
    relations = {}
    for idx, row in df.iterrows():
        manager = row.iloc[0]
        sub_cell = row.iloc[1]
        
        if pd.isna(manager) or pd.isna(sub_cell):
            continue
        
        manager = str(manager).strip()
        # 支持逗号分隔多个下属
        subs = [s.strip() for s in str(sub_cell).split(",") if s.strip()]
        
        if manager in relations:
            relations[manager].extend(subs)
        else:
            relations[manager] = subs
    
    # 打印读取结果
    for manager, subs in relations.items():
        print(f"  {manager} <- {subs}")
    
    print(f"\n共读取 {len(relations)} 组负责人关系")
    return relations


def get_manager_relations():
    """询问用户选择输入方式"""
    print()
    print("=" * 50)
    print("负责人-下属关系配置")
    print("=" * 50)
    print("说明: 下属的数据将加总到负责人数据上")
    print()
    print("请选择输入方式:")
    print("  1. 手动交互输入")
    print("  2. 从Excel文件读取")
    print("  0. 跳过（无需配置）")
    print()
    
    choice = input("请选择 (0/1/2): ").strip()
    
    if choice == '1':
        return input_manager_relations_interactive()
    elif choice == '2':
        return input_manager_relations_excel()
    else:
        return {}


def aggregate_data(name_to_row, relations, data_start_col):
    """
    汇总负责人数据：将下属的数值列加到负责人上
    返回: 汇总后的 name_to_row（已修改负责人数据）
    """
    for manager, subordinates in relations.items():
        if manager not in name_to_row:
            print(f"  警告: 负责人 '{manager}' 不在总文件中，跳过")
            continue
        
        manager_row = name_to_row[manager]
        
        for sub in subordinates:
            if sub not in name_to_row:
                print(f"  警告: 下属 '{sub}' 不在总文件中，跳过")
                continue
            
            sub_row = name_to_row[sub]
            
            # 从 data_start_col 开始，所有数值列加总
            for col_idx in range(data_start_col, len(manager_row)):
                m_val = manager_row[col_idx]
                s_val = sub_row[col_idx]
                
                if isinstance(m_val, (int, float)) and not pd.isna(m_val):
                    if isinstance(s_val, (int, float)) and not pd.isna(s_val):
                        manager_row[col_idx] = m_val + s_val
                elif isinstance(s_val, (int, float)) and not pd.isna(s_val):
                    manager_row[col_idx] = s_val
        
        name_to_row[manager] = manager_row
    
    return name_to_row


def generate_relation_report(name_to_row, relations, output_dir, data_start_col, total_file_path):
    """
    生成负责人-下属数据明细表
    格式: 负责人行（汇总后）+ 每个下属行（原始数据）
    """
    if not relations:
        return
    
    # 读取总文件前6行，合并第2-5行作为多级表头
    df_header = pd.read_excel(total_file_path, header=None, nrows=6)
    
    # 总表表头结构：第2行=大类，第3行=子类，第4行=指标名，第5行=单位/说明
    # 合并策略：从上往下找到非空值拼接
    num_cols = len(df_header.columns)
    total_headers = []
    
    for col_idx in range(num_cols):
        parts = []
        for row_idx in range(1, 5):  # 第2-5行（索引1-4）
            val = df_header.iloc[row_idx, col_idx]
            if pd.notna(val):
                text = str(val).strip()
                if text and text not in parts:
                    parts.append(text)
        if parts:
            total_headers.append("\n".join(parts))
        else:
            total_headers.append(f"列{col_idx}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "负责人汇总明细"
    
    header_font = Font(bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写表头
    headers = ["角色", "姓名"]
    sample_row = list(name_to_row.values())[0] if name_to_row else []
    for i in range(data_start_col, len(sample_row)):
        if i < len(total_headers):
            headers.append(total_headers[i])
        else:
            headers.append(f"列{i}")
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # 设置表头行高度
    ws.row_dimensions[1].height = 60
    
    current_row = 2
    
    for manager, subordinates in relations.items():
        if manager not in name_to_row:
            continue
        
        # 负责人行（汇总后的数据）
        manager_data = name_to_row[manager]
        ws.cell(row=current_row, column=1, value="负责人（汇总）")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=manager)
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        
        for i, col_idx in enumerate(range(data_start_col, len(manager_data))):
            val = manager_data[col_idx]
            if pd.notna(val):
                ws.cell(row=current_row, column=3 + i, value=val)
        current_row += 1
        
        # 下属行（原始数据保存在 relations 的上下文中）
        for sub in subordinates:
            if sub not in name_to_row:
                ws.cell(row=current_row, column=1, value="下属（未找到）")
                ws.cell(row=current_row, column=2, value=sub)
                current_row += 1
                continue
            
            sub_data = name_to_row[sub]
            ws.cell(row=current_row, column=1, value="  下属")
            ws.cell(row=current_row, column=2, value=sub)
            
            for i, col_idx in enumerate(range(data_start_col, len(sub_data))):
                val = sub_data[col_idx]
                if pd.notna(val):
                    ws.cell(row=current_row, column=3 + i, value=val)
            current_row += 1
        
        # 空行分隔
        current_row += 1
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    
    output_path = output_dir / "负责人汇总明细表.xlsx"
    wb.save(output_path)
    print(f"  负责人汇总明细表保存到: {output_path}")
    return output_path


def main():
    config = CONFIG
    output_dir = Path(config["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    
    args = sys.argv[1:]
    
    # 没有参数时，交互式询问
    if len(args) == 0:
        print("=" * 50)
        print("工资计算工具 v2")
        print("=" * 50)
        print()
        print("请选择运行模式：")
        print("  1. 使用内置模板（只需总文件）")
        print("  2. 使用自定义文件（总文件 + 自有表 + 外包表）")
        print()
        choice = input("请选择 (1/2): ").strip()
        print()
        
        if choice == '1':
            total_file = input("请输入总文件路径: ").strip().strip('"')
            mode = "template"
            print("模式: 使用内嵌模板（人员本月无变动）")
        elif choice == '2':
            total_file = input("请输入总文件路径: ").strip().strip('"')
            own_file = input("请输入自有表路径: ").strip().strip('"')
            outsource_file = input("请输入外包表路径: ").strip().strip('"')
            mode = "custom"
            print("模式: 使用自定义文件")
        else:
            print("无效选择，退出。")
            try:
                input("按回车键退出...")
            except:
                pass
            sys.exit(1)
    elif len(args) == 1:
        total_file = args[0]
        mode = "template"
        print("模式: 使用内嵌模板（人员本月无变动）")
    elif len(args) == 3:
        total_file, own_file, outsource_file = args
        mode = "custom"
        print("模式: 使用自定义文件")
    else:
        print("用法:")
        print("  拖拽文件到exe上运行，或使用命令行：")
        print("  工资计算工具v2.exe <总文件>                    # 使用内嵌模板")
        print("  工资计算工具v2.exe <总文件> <自有表> <外包表>   # 使用自定义文件")
        try:
            input("\n按回车键退出...")
        except:
            pass
        sys.exit(1)
    
    if not os.path.exists(total_file):
        print(f"错误: 总文件不存在 - {total_file}")
        try:
            input("\n按回车键退出...")
        except:
            pass
        sys.exit(1)
    
    print()
    print("=" * 50)
    print("工资计算工作流 v2")
    print("=" * 50)
    print()
    
    # 1. 加载总文件
    data_start_col = config["total"]["data_start_col"]
    name_to_row = load_total_file(total_file, config)
    
    # 2. 询问负责人-下属关系
    relations = get_manager_relations()
    
    # 3. 如果有关系，保存原始数据用于明细表，再汇总
    if relations:
        import copy
        name_to_row_original = copy.deepcopy(name_to_row)
        print()
        print("执行数据汇总...")
        name_to_row = aggregate_data(name_to_row, relations, data_start_col)
        print("  汇总完成")
    else:
        name_to_row_original = name_to_row
    
    # 映射规则
    own_mapping = {i: i - 7 for i in range(7, 43)}
    outsource_mapping = COLUMN_MAPPING_RULES["default"]
    
    print()
    if mode == "template":
        own_tmp = decode_template(OWN_TEMPLATE_B64)
        out_tmp = decode_template(OUTSOURCE_TEMPLATE_B64)
        
        try:
            print("处理自有揽投员核算表（模板）...")
            wb_own = load_workbook(own_tmp)
            own_names = set()
            
            for sheet_name, sheet_cfg in config["own"]["sheets"].items():
                paste_data(wb_own, sheet_name, sheet_cfg, name_to_row, own_mapping)
                ws = wb_own[sheet_name]
                for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                    if val:
                        own_names.add(str(val).strip())
            
            own_output = output_dir / "自有揽投员核算表标快.xlsx"
            wb_own.save(own_output)
            print(f"  保存到: {own_output}")
            print()
            
            print("处理外包人员标快（模板）...")
            wb_out = load_workbook(out_tmp)
            outsource_names = set()
            
            for sheet_name, sheet_cfg in config["outsource"]["sheets"].items():
                paste_data(wb_out, sheet_name, sheet_cfg, name_to_row, outsource_mapping)
                ws = wb_out[sheet_name]
                for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                    if val:
                        outsource_names.add(str(val).strip())
            
            outsource_output = output_dir / "外包人员标快计件薪酬核算表.xlsx"
            wb_out.save(outsource_output)
            print(f"  保存到: {outsource_output}")
        finally:
            os.unlink(own_tmp)
            os.unlink(out_tmp)
    
    else:
        print("处理自有揽投员核算表...")
        wb_own = load_workbook(own_file)
        own_names = set()
        
        for sheet_name, sheet_cfg in config["own_custom"]["sheets"].items():
            paste_data(wb_own, sheet_name, sheet_cfg, name_to_row, own_mapping)
            ws = wb_own[sheet_name]
            for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                if val:
                    own_names.add(str(val).strip())
        
        own_output = output_dir / Path(own_file).name
        wb_own.save(own_output)
        print(f"  保存到: {own_output}")
        print()
        
        print("处理外包人员标快...")
        wb_out = load_workbook(outsource_file)
        outsource_names = set()
        
        for sheet_name, sheet_cfg in config["outsource_custom"]["sheets"].items():
            paste_data(wb_out, sheet_name, sheet_cfg, name_to_row, outsource_mapping)
            ws = wb_out[sheet_name]
            for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                if val:
                    outsource_names.add(str(val).strip())
        
        outsource_output = output_dir / Path(outsource_file).name
        wb_out.save(outsource_output)
        print(f"  保存到: {outsource_output}")
    
    # 4. 生成负责人汇总明细表
    if relations:
        print()
        print("生成负责人汇总明细表...")
        generate_relation_report(name_to_row, relations, output_dir, data_start_col, total_file)
    
    # 汇总
    salesperson = set(name_to_row.keys()) - own_names - outsource_names
    
    print()
    print("=" * 50)
    print("汇总")
    print("=" * 50)
    print(f"平舆总人数: {len(name_to_row)}")
    print(f"自有: {len(own_names & set(name_to_row.keys()))}")
    print(f"外包: {len(outsource_names & set(name_to_row.keys()))}")
    print(f"营业员: {len(salesperson)}")
    if relations:
        print(f"负责人关系: {len(relations)} 组")
    print(f"\n输出目录: {output_dir}")
    print()
    try:
        input("按回车键退出...")
    except EOFError:
        pass


if __name__ == "__main__":
    main()
'''

output_path = os.path.join(base_dir, 'run_salary_v2.py')
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(script)

print(f"Generated: {output_path}")
print(f"File size: {os.path.getsize(output_path)} bytes")
