# -*- coding: utf-8 -*-
"""
工资计算工作流 - Step 1：从总文件匹配人员信息到自有表和外包表

v3: 使用硬编码分段映射，解决电商业务收入列不一致问题。
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ============================================================
# 配置区 - 每月修改这里的文件名
# ============================================================
CONFIG = {
    # 总文件（.xls 格式）- 每月更新文件名
    "total_file": r"C:\Users\yuson\Desktop\202605--揽投人员基础数据表.xls",
    
    # 自有揽投员核算表标快（.xlsx 格式）- 每月更新文件名
    "own_file": r"C:\Users\yuson\Desktop\2026年5月自有揽投员核算表标快.xlsx",
    
    # 外包人员标快（.xlsx 格式）- 每月更新文件名
    "outsource_file": r"C:\Users\yuson\Desktop\平舆外包人员标快2026年5月计件薪酬核算表.xlsx",
    
    # 输出目录
    "output_dir": r"C:\Users\yuson\.openclaw\workspace\salary-calculator\output",
    
    # 总文件结构参数
    "total": {
        "unit_col": 1,
        "name_col": 4,
        "data_start_col": 7,    # 数据从 col 7 开始（0-based索引）
        "data_start_row": 5,    # 数据从第 5 行开始（0-based索引）
        "filter_unit": "平舆",
    },
    
    # 自有表结构参数
    "own": {
        "sheets": {
            "县城自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
            "乡镇自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
        }
    },
    
    # 外包表结构参数
    "outsource": {
        "sheets": {
            "县城外包": {
                "name_col": 3,
                "data_start_row": 6,
                "data_start_col": 6,
            },
            "乡邮外包 ": {
                "name_col": 4,
                "data_start_row": 6,
                "data_start_col": 7,
            },
        }
    },
}


# 硬编码列映射规则（针对外包表）
# total_col -> target_offset (相对于target_start_col的偏移)
# 总文件列索引（0-based） -> 目标文件相对偏移
COLUMN_MAPPING_RULES = {
    # 县城外包和乡邮外包的映射规则
    "default": {
        # col 7-31: 1:1映射（退换货+自助计件+特快业务）
        **{i: i - 7 for i in range(7, 32)},
        # col 32-34: 电商业务量（1:1映射）
        32: 25,  # col32 -> offset 25 (target col 32 for 县城外包 data_start_col=6)
        33: 26,
        34: 27,
        # col 35: "散户:结算差额率30%以上" -> 跳过，不映射
        # col 36-40: 电商业务收入（映射到 target offset 28-32）
        36: 28,  # 结算差额率30%以上
        37: 29,  # 25%-30%
        38: 30,  # 20%-25%
        39: 31,  # 15%-20%
        40: 32,  # 低于15%不核发
        41: 33,  # 计酬
        # col 42-44: 合计、全表合计、校对（通常不需要粘贴）
    },
}


def load_total_file(config):
    """读取总文件，筛选平舆人员，返回{姓名: 整行list}"""
    total_cfg = config["total"]
    df = pd.read_excel(config["total_file"], header=None)
    
    mask = df.iloc[:, total_cfg["unit_col"]] == total_cfg["filter_unit"]
    filtered = df[mask]
    
    name_to_row = {}
    for idx, row in filtered.iterrows():
        name = row.iloc[total_cfg["name_col"]]
        if pd.notna(name):
            name_to_row[str(name).strip()] = row.tolist()
    
    print(f"总文件筛选出 {total_cfg['filter_unit']} 人员: {len(name_to_row)} 人")
    return name_to_row


def paste_data_with_hardcoded_mapping(wb, sheet_name, sheet_cfg, name_to_row, mapping_rule):
    """使用硬编码映射粘贴数据"""
    ws = wb[sheet_name]
    name_col = sheet_cfg["name_col"]
    data_start_col = sheet_cfg["data_start_col"]
    data_start_row = sheet_cfg["data_start_row"]
    
    matched = 0
    not_found = []
    
    # 遍历目标sheet中的人员行
    for row_idx in range(data_start_row + 1, ws.max_row + 1):  # openpyxl 1-based
        cell_value = ws.cell(row=row_idx, column=name_col + 1).value
        if cell_value is None:
            continue
        
        name = str(cell_value).strip()
        if name in name_to_row:
            row_data = name_to_row[name]
            # 按硬编码映射粘贴
            for total_col_idx, target_offset in mapping_rule.items():
                if total_col_idx < len(row_data):
                    val = row_data[total_col_idx]
                    if pd.notna(val):
                        target_col = data_start_col + 1 + target_offset  # openpyxl 1-based
                        ws.cell(row=row_idx, column=target_col, value=val)
            matched += 1
        else:
            not_found.append(name)
    
    print(f"  [{sheet_name}] 匹配成功: {matched} 人", end="")
    if not_found:
        print(f", 未在总文件中找到: {not_found}")
    else:
        print()
    
    return matched


def create_other_outsource_sheet(wb, other_names, name_to_row):
    """在外包表中新建"其他外包"sheet"""
    if not other_names:
        return
    
    if "其他外包" in wb.sheetnames:
        del wb["其他外包"]
    ws = wb.create_sheet("其他外包")
    
    headers = ["单位", "网点名称", "人员编码", "姓名", "用工类型", "岗位类型"]
    for i, h in enumerate(headers):
        ws.cell(row=1, column=i + 1, value=h)
    
    row_num = 2
    for name in sorted(other_names):
        if name in name_to_row:
            row_data = name_to_row[name]
            ws.cell(row=row_num, column=1, value=row_data[1])
            ws.cell(row=row_num, column=2, value=row_data[2])
            ws.cell(row=row_num, column=3, value=row_data[3])
            ws.cell(row=row_num, column=4, value=name)
            ws.cell(row=row_num, column=5, value=row_data[5])
            ws.cell(row=row_num, column=6, value=row_data[6])
            row_num += 1
    
    print(f"  [其他外包] 新建sheet，写入 {row_num - 2} 人")


def main():
    config = CONFIG
    output_dir = Path(config["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("=" * 50)
    print("工资计算工作流 - Step 1: 人员信息匹配")
    print("=" * 50)
    print()
    
    # 1. 加载总文件
    name_to_row = load_total_file(config)
    print()
    
    # 2. 处理自有表（自有表结构和总文件一致，1:1映射）
    print("处理自有揽投员核算表...")
    own_file = config["own_file"]
    wb_own = load_workbook(own_file)
    own_names = set()
    
    # 自有表用简单1:1映射（col 7-44 -> offset 0-37）
    own_mapping = {i: i - 7 for i in range(7, 45)}
    
    for sheet_name, sheet_cfg in config["own"]["sheets"].items():
        paste_data_with_hardcoded_mapping(wb_own, sheet_name, sheet_cfg, name_to_row, own_mapping)
        ws = wb_own[sheet_name]
        for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
            if val:
                own_names.add(str(val).strip())
    
    own_output = output_dir / Path(own_file).name
    wb_own.save(own_output)
    print(f"  保存到: {own_output}")
    print()
    
    # 3. 处理外包表（用硬编码映射）
    print("处理外包人员标快...")
    outsource_file = config["outsource_file"]
    wb_out = load_workbook(outsource_file)
    outsource_names = set()
    
    outsource_mapping = COLUMN_MAPPING_RULES["default"]
    
    for sheet_name, sheet_cfg in config["outsource"]["sheets"].items():
        paste_data_with_hardcoded_mapping(wb_out, sheet_name, sheet_cfg, name_to_row, outsource_mapping)
        ws = wb_out[sheet_name]
        for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
            if val:
                outsource_names.add(str(val).strip())
    
    # 4. 识别营业员
    salesperson = set(name_to_row.keys()) - own_names - outsource_names
    print(f"\n营业员（自动识别）: {len(salesperson)} 人")
    
    # 5. 其他外包
    all_outsource = set(name_to_row.keys()) - own_names - salesperson
    other_outsource = all_outsource - outsource_names
    
    if other_outsource:
        print(f"\n其他外包人员: {len(other_outsource)} 人")
        create_other_outsource_sheet(wb_out, other_outsource, name_to_row)
    else:
        print("\n无其他外包人员")
    
    outsource_output = output_dir / Path(outsource_file).name
    wb_out.save(outsource_output)
    print(f"  保存到: {outsource_output}")
    
    # 6. 汇总
    print("\n" + "=" * 50)
    print("汇总")
    print("=" * 50)
    print(f"平舆总人数: {len(name_to_row)}")
    print(f"自有: {len(own_names & set(name_to_row.keys()))}")
    print(f"外包(已有sheet): {len(outsource_names & set(name_to_row.keys()))}")
    print(f"其他外包: {len(other_outsource)}")
    print(f"营业员: {len(salesperson)}")
    print(f"\n输出目录: {output_dir}")


if __name__ == "__main__":
    main()
