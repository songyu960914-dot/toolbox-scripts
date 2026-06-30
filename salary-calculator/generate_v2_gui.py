# -*- coding: utf-8 -*-
"""生成 run_salary_v2_gui.py - 带GUI界面的版本"""
import os

base_dir = r'C:\Users\yuson\.openclaw\workspace\salary-calculator'
own_b64 = open(os.path.join(base_dir, 'own_template_b64.txt')).read().strip()  # 需要本地生成
out_b64 = open(os.path.join(base_dir, 'outsource_template_b64.txt')).read().strip()  # 需要本地生成

script = r'''# -*- coding: utf-8 -*-
"""
工资计算工具 v2 - GUI版本
"""

import os
import sys
import copy
import base64
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# ============================================================
# 内嵌模板（base64编码）
# ============================================================
OWN_TEMPLATE_B64 = """
''' + own_b64 + r'''
"""

OUTSOURCE_TEMPLATE_B64 = """
''' + out_b64 + r'''
"""


# ============================================================
# 配置区
# ============================================================
CONFIG = {
    "output_dir": os.path.join(os.path.expanduser("~"), "Desktop", "工资计算结果"),
    "total": {
        "unit_col": 1,
        "name_col": 4,
        "data_start_col": 7,
        "data_start_row": 5,
        "filter_unit": "平舆",
    },
    "own": {
        "sheets": {
            "县城自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
            "乡镇自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
        }
    },
    "outsource": {
        "sheets": {
            "县城外包": {
                "name_col": 3,
                "data_start_row": 6,
                "data_start_col": 6,
            },
            "乡邮外包": {
                "name_col": 4,
                "data_start_row": 6,
                "data_start_col": 7,
            },
        }
    },
    "own_custom": {
        "sheets": {
            "县城自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
            "乡镇自有人员新政策": {
                "name_col": 3,
                "data_start_row": 4,
                "data_start_col": 6,
            },
        }
    },
    "outsource_custom": {
        "sheets": {
            "县城外包": {
                "name_col": 3,
                "data_start_row": 6,
                "data_start_col": 6,
            },
            "乡邮外包 ": {
                "name_col": 4,
                "data_start_row": 6,
                "data_start_col": 7,
            },
        }
    },
}

COLUMN_MAPPING_RULES = {
    "default": {
        **{i: i - 7 for i in range(7, 32)},
        32: 25, 33: 26, 34: 27,
        36: 28, 37: 29, 38: 30, 39: 31, 40: 32, 41: 33,
    },
}


# ============================================================
# 核心逻辑（复用命令行版）
# ============================================================
def decode_template(b64_str):
    data = base64.b64decode(b64_str.strip())
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.write(data)
    tmp.close()
    return tmp.name


def load_total_file(total_file, config):
    total_cfg = config["total"]
    df = pd.read_excel(total_file, header=None)
    mask = df.iloc[:, total_cfg["unit_col"]] == total_cfg["filter_unit"]
    filtered = df[mask]
    name_to_row = {}
    for idx, row in filtered.iterrows():
        name = row.iloc[total_cfg["name_col"]]
        if pd.notna(name):
            name_to_row[str(name).strip()] = row.tolist()
    return name_to_row


def paste_data(wb, sheet_name, sheet_cfg, name_to_row, mapping_rule):
    ws = wb[sheet_name]
    name_col = sheet_cfg["name_col"]
    data_start_col = sheet_cfg["data_start_col"]
    data_start_row = sheet_cfg["data_start_row"]
    matched = 0
    not_found = []
    for row_idx in range(data_start_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=name_col + 1).value
        if cell_value is None:
            continue
        name = str(cell_value).strip()
        if name in name_to_row:
            row_data = name_to_row[name]
            for total_col_idx, target_offset in mapping_rule.items():
                if total_col_idx < len(row_data):
                    val = row_data[total_col_idx]
                    if pd.notna(val):
                        target_col = data_start_col + 1 + target_offset
                        ws.cell(row=row_idx, column=target_col, value=val)
            matched += 1
        else:
            not_found.append(name)
    return matched, not_found


def aggregate_data(name_to_row, relations, data_start_col):
    warnings = []
    for manager, subordinates in relations.items():
        if manager not in name_to_row:
            warnings.append(f"负责人 '{manager}' 不在总文件中，跳过")
            continue
        manager_row = name_to_row[manager]
        for sub in subordinates:
            if sub not in name_to_row:
                warnings.append(f"下属 '{sub}' 不在总文件中，跳过")
                continue
            sub_row = name_to_row[sub]
            for col_idx in range(data_start_col, len(manager_row)):
                m_val = manager_row[col_idx]
                s_val = sub_row[col_idx]
                if isinstance(m_val, (int, float)) and not pd.isna(m_val):
                    if isinstance(s_val, (int, float)) and not pd.isna(s_val):
                        manager_row[col_idx] = m_val + s_val
                elif isinstance(s_val, (int, float)) and not pd.isna(s_val):
                    manager_row[col_idx] = s_val
        name_to_row[manager] = manager_row
    return name_to_row, warnings


def generate_relation_report(name_to_row, relations, output_dir, data_start_col, total_file_path):
    if not relations:
        return None
    
    # 读取总文件前6行，合并第2-5行作为多级表头
    df_header = pd.read_excel(total_file_path, header=None, nrows=6)
    
    num_cols = len(df_header.columns)
    total_headers = []
    for col_idx in range(num_cols):
        parts = []
        for row_idx in range(1, 5):  # 第2-5行（索引1-4）
            val = df_header.iloc[row_idx, col_idx]
            if pd.notna(val):
                text = str(val).strip()
                if text and text not in parts:
                    parts.append(text)
        if parts:
            total_headers.append("\n".join(parts))
        else:
            total_headers.append(f"列{col_idx}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "负责人汇总明细"
    
    header_font = Font(bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["角色", "姓名"]
    sample_row = list(name_to_row.values())[0] if name_to_row else []
    for i in range(data_start_col, len(sample_row)):
        if i < len(total_headers):
            headers.append(total_headers[i])
        else:
            headers.append(f"列{i}")
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 60
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 60
    
    current_row = 2
    for manager, subordinates in relations.items():
        if manager not in name_to_row:
            continue
        manager_data = name_to_row[manager]
        ws.cell(row=current_row, column=1, value="负责人（汇总）")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=manager)
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        for i, col_idx in enumerate(range(data_start_col, len(manager_data))):
            val = manager_data[col_idx]
            if pd.notna(val):
                ws.cell(row=current_row, column=3 + i, value=val)
        current_row += 1
        
        for sub in subordinates:
            if sub not in name_to_row:
                ws.cell(row=current_row, column=1, value="下属（未找到）")
                ws.cell(row=current_row, column=2, value=sub)
                current_row += 1
                continue
            sub_data = name_to_row[sub]
            ws.cell(row=current_row, column=1, value="  下属")
            ws.cell(row=current_row, column=2, value=sub)
            for i, col_idx in enumerate(range(data_start_col, len(sub_data))):
                val = sub_data[col_idx]
                if pd.notna(val):
                    ws.cell(row=current_row, column=3 + i, value=val)
            current_row += 1
        current_row += 1
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    
    output_path = output_dir / "负责人汇总明细表.xlsx"
    wb.save(output_path)
    return output_path


# ============================================================
# GUI 主界面
# ============================================================
class SalaryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("工资计算工具 v2")
        self.root.geometry("700x650")
        self.root.resizable(True, True)
        
        # 变量
        self.mode_var = tk.StringVar(value="template")
        self.total_file_var = tk.StringVar()
        self.own_file_var = tk.StringVar()
        self.outsource_file_var = tk.StringVar()
        self.relations = {}
        
        self.build_ui()
    
    def build_ui(self):
        # 标题
        title_frame = tk.Frame(self.root, pady=10)
        title_frame.pack(fill='x')
        tk.Label(title_frame, text="工资计算工具 v2", font=("微软雅黑", 16, "bold")).pack()
        
        # 模式选择
        mode_frame = tk.LabelFrame(self.root, text="运行模式", padx=10, pady=5)
        mode_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Radiobutton(mode_frame, text="使用内置模板（人员无变动）", 
                       variable=self.mode_var, value="template",
                       command=self.toggle_mode).pack(anchor='w')
        tk.Radiobutton(mode_frame, text="使用自定义文件（人员有变动）", 
                       variable=self.mode_var, value="custom",
                       command=self.toggle_mode).pack(anchor='w')
        
        # 文件选择
        file_frame = tk.LabelFrame(self.root, text="文件选择", padx=10, pady=5)
        file_frame.pack(fill='x', padx=10, pady=5)
        
        # 总文件
        row1 = tk.Frame(file_frame)
        row1.pack(fill='x', pady=2)
        tk.Label(row1, text="总文件:", width=8, anchor='w').pack(side='left')
        tk.Entry(row1, textvariable=self.total_file_var, width=50).pack(side='left', fill='x', expand=True)
        tk.Button(row1, text="浏览", command=self.browse_total).pack(side='left', padx=5)
        
        # 自有表
        row2 = tk.Frame(file_frame)
        row2.pack(fill='x', pady=2)
        tk.Label(row2, text="自有表:", width=8, anchor='w').pack(side='left')
        self.own_entry = tk.Entry(row2, textvariable=self.own_file_var, width=50, state='disabled')
        self.own_entry.pack(side='left', fill='x', expand=True)
        self.own_btn = tk.Button(row2, text="浏览", command=self.browse_own, state='disabled')
        self.own_btn.pack(side='left', padx=5)
        
        # 外包表
        row3 = tk.Frame(file_frame)
        row3.pack(fill='x', pady=2)
        tk.Label(row3, text="外包表:", width=8, anchor='w').pack(side='left')
        self.out_entry = tk.Entry(row3, textvariable=self.outsource_file_var, width=50, state='disabled')
        self.out_entry.pack(side='left', fill='x', expand=True)
        self.out_btn = tk.Button(row3, text="浏览", command=self.browse_outsource, state='disabled')
        self.out_btn.pack(side='left', padx=5)
        
        # 负责人-下属关系
        relation_frame = tk.LabelFrame(self.root, text="负责人-下属关系（可选）", padx=10, pady=5)
        relation_frame.pack(fill='x', padx=10, pady=5)
        
        btn_row = tk.Frame(relation_frame)
        btn_row.pack(fill='x', pady=2)
        tk.Button(btn_row, text="手动添加", command=self.add_relation).pack(side='left', padx=5)
        tk.Button(btn_row, text="从Excel导入", command=self.import_relations).pack(side='left', padx=5)
        tk.Button(btn_row, text="清空", command=self.clear_relations).pack(side='left', padx=5)
        
        self.relation_text = tk.Text(relation_frame, height=4, width=60, state='disabled')
        self.relation_text.pack(fill='x', pady=2)
        
        # 执行按钮
        btn_frame = tk.Frame(self.root, pady=10)
        btn_frame.pack(fill='x')
        self.run_btn = tk.Button(btn_frame, text="开始计算", font=("微软雅黑", 12, "bold"),
                                 bg="#4CAF50", fg="white", padx=20, pady=5,
                                 command=self.run_calculation)
        self.run_btn.pack()
        
        # 日志输出
        log_frame = tk.LabelFrame(self.root, text="运行日志", padx=10, pady=5)
        log_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, width=60)
        self.log_text.pack(fill='both', expand=True)
    
    def toggle_mode(self):
        if self.mode_var.get() == "custom":
            self.own_entry.config(state='normal')
            self.own_btn.config(state='normal')
            self.out_entry.config(state='normal')
            self.out_btn.config(state='normal')
        else:
            self.own_entry.config(state='disabled')
            self.own_btn.config(state='disabled')
            self.out_entry.config(state='disabled')
            self.out_btn.config(state='disabled')
    
    def browse_total(self):
        f = filedialog.askopenfilename(title="选择总文件", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if f:
            self.total_file_var.set(f)
    
    def browse_own(self):
        f = filedialog.askopenfilename(title="选择自有表", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if f:
            self.own_file_var.set(f)
    
    def browse_outsource(self):
        f = filedialog.askopenfilename(title="选择外包表", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if f:
            self.outsource_file_var.set(f)
    
    def add_relation(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("添加负责人-下属关系")
        dialog.geometry("400x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="负责人姓名:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        manager_entry = tk.Entry(dialog, width=30)
        manager_entry.grid(row=0, column=1, padx=10, pady=10)
        
        tk.Label(dialog, text="下属姓名:").grid(row=1, column=0, padx=10, pady=10, sticky='w')
        sub_entry = tk.Entry(dialog, width=30)
        sub_entry.grid(row=1, column=1, padx=10, pady=10)
        tk.Label(dialog, text="（多人用逗号分隔）").grid(row=2, column=1, sticky='w', padx=10)
        
        def confirm():
            manager = manager_entry.get().strip()
            subs_str = sub_entry.get().strip()
            if manager and subs_str:
                subs = [s.strip() for s in subs_str.split(",") if s.strip()]
                if manager in self.relations:
                    self.relations[manager].extend(subs)
                else:
                    self.relations[manager] = subs
                self.update_relation_display()
            dialog.destroy()
        
        tk.Button(dialog, text="确认", command=confirm).grid(row=3, column=1, pady=10)
    
    def import_relations(self):
        f = filedialog.askopenfilename(title="选择关系表", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if not f:
            return
        try:
            df = pd.read_excel(f, header=None)
            for idx, row in df.iterrows():
                manager = row.iloc[0]
                sub_cell = row.iloc[1]
                if pd.isna(manager) or pd.isna(sub_cell):
                    continue
                manager = str(manager).strip()
                subs = [s.strip() for s in str(sub_cell).split(",") if s.strip()]
                if manager in self.relations:
                    self.relations[manager].extend(subs)
                else:
                    self.relations[manager] = subs
            self.update_relation_display()
            messagebox.showinfo("导入成功", f"已导入 {len(self.relations)} 组负责人关系")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))
    
    def clear_relations(self):
        self.relations = {}
        self.update_relation_display()
    
    def update_relation_display(self):
        self.relation_text.config(state='normal')
        self.relation_text.delete('1.0', 'end')
        if self.relations:
            for manager, subs in self.relations.items():
                self.relation_text.insert('end', f"{manager} ← {', '.join(subs)}\n")
        else:
            self.relation_text.insert('end', "（未配置）")
        self.relation_text.config(state='disabled')
    
    def log(self, msg):
        self.log_text.insert('end', msg + '\n')
        self.log_text.see('end')
        self.root.update_idletasks()
    
    def run_calculation(self):
        # 验证输入
        total_file = self.total_file_var.get().strip()
        if not total_file:
            messagebox.showwarning("提示", "请选择总文件")
            return
        if not os.path.exists(total_file):
            messagebox.showerror("错误", f"总文件不存在: {total_file}")
            return
        
        mode = self.mode_var.get()
        if mode == "custom":
            own_file = self.own_file_var.get().strip()
            outsource_file = self.outsource_file_var.get().strip()
            if not own_file or not outsource_file:
                messagebox.showwarning("提示", "请选择自有表和外包表")
                return
        
        # 禁用按钮防止重复点击
        self.run_btn.config(state='disabled', text="计算中...")
        self.log_text.delete('1.0', 'end')
        
        # 在线程中运行避免界面卡顿
        thread = threading.Thread(target=self._do_calculation, args=(total_file, mode))
        thread.start()
    
    def _do_calculation(self, total_file, mode):
        try:
            config = CONFIG
            output_dir = Path(config["output_dir"])
            output_dir.mkdir(parents=True, exist_ok=True)
            data_start_col = config["total"]["data_start_col"]
            
            # 1. 加载总文件
            self.log("正在读取总文件...")
            name_to_row = load_total_file(total_file, config)
            self.log(f"筛选出平舆人员: {len(name_to_row)} 人")
            
            # 2. 负责人关系处理
            relations = self.relations
            if relations:
                name_to_row_original = copy.deepcopy(name_to_row)
                self.log(f"\n执行负责人数据汇总（{len(relations)} 组）...")
                name_to_row, warnings = aggregate_data(name_to_row, relations, data_start_col)
                for w in warnings:
                    self.log(f"  警告: {w}")
                self.log("  汇总完成")
            
            # 映射规则
            own_mapping = {i: i - 7 for i in range(7, 43)}
            outsource_mapping = COLUMN_MAPPING_RULES["default"]
            
            if mode == "template":
                own_tmp = decode_template(OWN_TEMPLATE_B64)
                out_tmp = decode_template(OUTSOURCE_TEMPLATE_B64)
                
                try:
                    self.log("\n处理自有揽投员核算表...")
                    wb_own = load_workbook(own_tmp)
                    own_names = set()
                    for sheet_name, sheet_cfg in config["own"]["sheets"].items():
                        matched, not_found = paste_data(wb_own, sheet_name, sheet_cfg, name_to_row, own_mapping)
                        ws = wb_own[sheet_name]
                        for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                            val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                            if val:
                                own_names.add(str(val).strip())
                        self.log(f"  [{sheet_name}] 匹配: {matched} 人")
                        if not_found:
                            self.log(f"    未找到: {not_found}")
                    
                    own_output = output_dir / "自有揽投员核算表标快.xlsx"
                    wb_own.save(own_output)
                    self.log(f"  保存: {own_output}")
                    
                    self.log("\n处理外包人员核算表...")
                    wb_out = load_workbook(out_tmp)
                    outsource_names = set()
                    for sheet_name, sheet_cfg in config["outsource"]["sheets"].items():
                        matched, not_found = paste_data(wb_out, sheet_name, sheet_cfg, name_to_row, outsource_mapping)
                        ws = wb_out[sheet_name]
                        for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                            val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                            if val:
                                outsource_names.add(str(val).strip())
                        self.log(f"  [{sheet_name}] 匹配: {matched} 人")
                        if not_found:
                            self.log(f"    未找到: {not_found}")
                    
                    outsource_output = output_dir / "外包人员标快计件薪酬核算表.xlsx"
                    wb_out.save(outsource_output)
                    self.log(f"  保存: {outsource_output}")
                finally:
                    os.unlink(own_tmp)
                    os.unlink(out_tmp)
            
            else:
                own_file = self.own_file_var.get().strip()
                outsource_file = self.outsource_file_var.get().strip()
                
                self.log("\n处理自有揽投员核算表...")
                wb_own = load_workbook(own_file)
                own_names = set()
                for sheet_name, sheet_cfg in config["own_custom"]["sheets"].items():
                    matched, not_found = paste_data(wb_own, sheet_name, sheet_cfg, name_to_row, own_mapping)
                    ws = wb_own[sheet_name]
                    for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                        val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                        if val:
                            own_names.add(str(val).strip())
                    self.log(f"  [{sheet_name}] 匹配: {matched} 人")
                    if not_found:
                        self.log(f"    未找到: {not_found}")
                
                own_output = output_dir / Path(own_file).name
                wb_own.save(own_output)
                self.log(f"  保存: {own_output}")
                
                self.log("\n处理外包人员核算表...")
                wb_out = load_workbook(outsource_file)
                outsource_names = set()
                for sheet_name, sheet_cfg in config["outsource_custom"]["sheets"].items():
                    matched, not_found = paste_data(wb_out, sheet_name, sheet_cfg, name_to_row, outsource_mapping)
                    ws = wb_out[sheet_name]
                    for row_idx in range(sheet_cfg["data_start_row"] + 1, ws.max_row + 1):
                        val = ws.cell(row=row_idx, column=sheet_cfg["name_col"] + 1).value
                        if val:
                            outsource_names.add(str(val).strip())
                    self.log(f"  [{sheet_name}] 匹配: {matched} 人")
                    if not_found:
                        self.log(f"    未找到: {not_found}")
                
                outsource_output = output_dir / Path(outsource_file).name
                wb_out.save(outsource_output)
                self.log(f"  保存: {outsource_output}")
            
            # 负责人汇总明细表
            if relations:
                self.log("\n生成负责人汇总明细表...")
                report_path = generate_relation_report(name_to_row, relations, output_dir, data_start_col, total_file)
                if report_path:
                    self.log(f"  保存: {report_path}")
            
            # 汇总
            salesperson = set(name_to_row.keys()) - own_names - outsource_names
            self.log("\n" + "=" * 40)
            self.log(f"平舆总人数: {len(name_to_row)}")
            self.log(f"自有: {len(own_names & set(name_to_row.keys()))}")
            self.log(f"外包: {len(outsource_names & set(name_to_row.keys()))}")
            self.log(f"营业员: {len(salesperson)}")
            self.log(f"\n输出目录: {output_dir}")
            self.log("\n✅ 计算完成！")
            
            # 打开输出目录
            os.startfile(str(output_dir))
            
        except PermissionError:
            self.log("\n❌ 错误: 输出文件被占用，请关闭已打开的Excel文件后重试")
        except Exception as e:
            self.log(f"\n❌ 错误: {e}")
        finally:
            self.root.after(0, lambda: self.run_btn.config(state='normal', text="开始计算"))


def main():
    root = tk.Tk()
    app = SalaryApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
'''

output_path = os.path.join(base_dir, 'run_salary_v2_gui.py')
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(script)

print(f"Generated: {output_path}")
print(f"File size: {os.path.getsize(output_path)} bytes")
